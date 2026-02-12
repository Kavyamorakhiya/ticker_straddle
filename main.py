import csv
import io
import math
import re
import calendar
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Dict, Optional, Set, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ======= CONFIG =======
EXCEL_PATH = "C:\\Users\\kavya\\Desktop\\kavya.xlsx"
SHEET_NAME = "Sheet1"
EXCHANGE_PREFIX = "NSE:"  # your sheet uses values like "NSE:RELIANCE"

# Index symbols we will treat as "indices" for NSE endpoints
INDEX_SYMBOLS = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTYIT", "NIFTYFINSERVICE"}

# NSE endpoints (public)
NSE_OC_INDEX_URL = "https://www.nseindia.com/api/option-chain-indices?symbol={sym}"
NSE_OC_EQUITY_URL = "https://www.nseindia.com/api/option-chain-equities?symbol={sym}"
NSE_BAN_CSV_URL   = "https://nsearchives.nseindia.com/content/fo/fo_secban.csv"
NSE_HOLIDAYS_URL  = "https://www.nseindia.com/api/holiday-master?type=trading"

UA_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
    "Referer": "https://www.nseindia.com/"
}

# Expiry changeover: last Thursday -> last Tuesday (contracts expiring on/after this series)
CHANGEOVER_DATE = date(2025, 9, 1)

# Row colors
FILL_RED    = PatternFill(start_color="FFF1C7C7", end_color="FFF1C7C7", fill_type="solid")   # banned
FILL_ORANGE = PatternFill(start_color="FFFCD8B7", end_color="FFFCD8B7", fill_type="solid")   # >=90 MWPL
FILL_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")   # 80-90 MWPL

# ======= HTTP helper =======

# ---- PATCH: robust NSE session + fetch ----
import random
_UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36",
]

class NSEClient:
    BASE = "https://www.nseindia.com"
    OC_PAGE = BASE + "/option-chain"
    # Endpoints for getting contract info (expiry dates)
    CONTRACT_INFO_IDX = BASE + "/api/option-chain-contract-info?symbol={sym}"
    CONTRACT_INFO_EQ  = BASE + "/api/option-chain-contract-info?symbol={sym}"
    # Using v3 API endpoints (works better than the old ones)
    OC_IDX_DATA = BASE + "/api/option-chain-v3?type=Indices&symbol={sym}&expiry={exp}"
    OC_EQ_DATA  = BASE + "/api/option-chain-v3?type=Equity&symbol={sym}&expiry={exp}"

    def __init__(self):
        self.s = requests.Session()
        self.cookies = {}
        self._warm()

    def _warm(self):
        """Warm up session by visiting the option-chain page and getting cookies."""
        # Use a fixed user agent that NSE accepts (from VarunS2002's solution)
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
            'accept-language': 'en,gu;q=0.9,hi;q=0.8',
            'accept-encoding': 'gzip, deflate, br'
        }
        self.s.headers.update(headers)
        
        try:
            print("[NSE WARMUP] Visiting option-chain page...")
            r = self.s.get(self.OC_PAGE, headers=headers, timeout=10)
            self.cookies = dict(self.s.cookies)
            print(f"[NSE WARMUP] Status: {r.status_code}")
            print(f"[NSE WARMUP] Cookies obtained: {list(self.cookies.keys())}")
            print("[NSE WARMUP] Session ready!")
        except Exception as e:
            print(f"[NSE WARMUP ERROR] Failed: {e}")
            raise

    def _get_json(self, url, tries=2, sleep_sec=1.5, symbol="unknown", timeout=10):
        """Fetch JSON from NSE API with explicit cookie handling."""
        for attempt in range(1, tries + 1):
            try:
                print(f"[NSE ATTEMPT {attempt}/{tries}] {symbol}")
                resp = self.s.get(url, cookies=self.cookies, timeout=timeout)
                print(f"[NSE RESPONSE] Status: {resp.status_code} | {symbol}")
                
                if resp.status_code == 200:
                    json_data = resp.json()
                    if json_data:  # Check if not empty
                        print(f"[NSE SUCCESS] Got data for {symbol}")
                        return json_data
                    else:
                        print(f"[NSE WARNING] Empty JSON for {symbol}, retrying...")
                        time.sleep(sleep_sec)
                        continue
                elif resp.status_code in (401, 403):
                    print(f"[NSE AUTH ERROR] {resp.status_code} - Re-warming session...")
                    time.sleep(sleep_sec)
                    self._warm()  # Refresh cookies
                    continue
                else:
                    print(f"[NSE ERROR] Status {resp.status_code} - retrying...")
                    time.sleep(sleep_sec)
                    continue
            except Exception as e:
                print(f"[NSE EXCEPTION] Attempt {attempt}: {str(e)[:100]}")
                time.sleep(sleep_sec)
                continue
        
        print(f"[NSE FINAL ERROR] All {tries} attempts failed for {symbol}")
        return {}

    def get_expiry_dates(self, symbol_clean: str, is_index: bool):
        """Get available expiry dates for a symbol."""
        sym_enc = requests.utils.quote(symbol_clean.upper())
        url = self.CONTRACT_INFO_IDX.format(sym=sym_enc) if is_index else self.CONTRACT_INFO_EQ.format(sym=sym_enc)
        return self._get_json(url, symbol=symbol_clean)

    def option_chain(self, symbol_clean: str, is_index: bool, expiry_date: str):
        """Fetch option chain data using v3 API with expiry date."""
        sym_enc = requests.utils.quote(symbol_clean.upper())
        exp_enc = requests.utils.quote(expiry_date)
        
        if is_index:
            url = self.OC_IDX_DATA.format(sym=sym_enc, exp=exp_enc)
        else:
            url = self.OC_EQ_DATA.format(sym=sym_enc, exp=exp_enc)
        
        return self._get_json(url, symbol=symbol_clean)

_nse = None
def make_session():
    global _nse
    if _nse is None:
        _nse = NSEClient()
    return _nse.s

def fetch_oc_dates(symbol_clean: str, is_index: bool) -> dict:
    """Fetch available expiry dates for a symbol."""
    global _nse
    if _nse is None:
        _nse = NSEClient()
    return _nse.get_expiry_dates(symbol_clean, is_index)

def fetch_oc_json(symbol_clean: str, is_index: bool, expiry_date: str) -> dict:
    """Fetch option chain JSON data for a given symbol and expiry date."""
    global _nse
    if _nse is None:
        _nse = NSEClient()
    return _nse.option_chain(symbol_clean, is_index, expiry_date)
# ---- END PATCH ----

# ======= Holidays =======
_holidays_cache: Optional[Set[date]] = None

def parse_date_flex(s: str) -> Optional[date]:
    s = (s or "").strip()
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    m = re.search(r"(\d{1,2})[-/\s]([A-Za-z]{3})[-/\s](\d{4})", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)}-{m.group(2)}-{m.group(3)}", "%d-%b-%Y").date()
        except Exception:
            return None
    return None

def get_trading_holidays() -> Set[date]:
    global _holidays_cache
    if _holidays_cache is not None:
        return _holidays_cache
    try:
        s = make_session()
        j = s.get(NSE_HOLIDAYS_URL, timeout=10).json()
        hols: Set[date] = set()
        for seg_key in ("FO", "TRADING", "CM"):
            for item in j.get(seg_key, []):
                dt_txt = item.get("tradingDate") or item.get("date")
                d = parse_date_flex(dt_txt)
                if d:
                    hols.add(d)
        _holidays_cache = hols
        return hols
    except Exception:
        _holidays_cache = set()
        return _holidays_cache

# ======= Expiry helpers (monthly only) =======
def _last_weekday_of_month(d: date, weekday: int) -> date:
    last_day = calendar.monthrange(d.year, d.month)[1]
    last_dt = date(d.year, d.month, last_day)
    delta = (last_dt.weekday() - weekday) % 7
    return last_dt - timedelta(days=delta)

def _prev_trading_day(dt: date, holidays: Set[date]) -> date:
    cur = dt
    while cur.weekday() >= 5 or cur in holidays:  # weekend or holiday
        cur -= timedelta(days=1)
    return cur

def last_thursday_monthly(d: date, holidays: Set[date]) -> date:
    return _prev_trading_day(_last_weekday_of_month(d, 3), holidays)  # Thu=3

def last_tuesday_monthly(d: date, holidays: Set[date]) -> date:
    return _prev_trading_day(_last_weekday_of_month(d, 1), holidays)  # Tue=1

def monthly_expiry_for_series(today: date) -> date:
    """Return the active monthly expiry date respecting Tue/Thu changeover + holidays."""
    holidays = get_trading_holidays()
    th_this = last_thursday_monthly(today, holidays)
    series_month = today if today <= th_this else date(today.year + (1 if today.month == 12 else 0),
                                                       1 if today.month == 12 else today.month + 1, 1)
    th_exp = last_thursday_monthly(series_month, holidays)
    tu_exp = last_tuesday_monthly(series_month, holidays)
    use_tuesday = (tu_exp >= CHANGEOVER_DATE)
    expiry = tu_exp if use_tuesday else th_exp
    if expiry < today:
        nm = date(series_month.year + (1 if series_month.month == 12 else 0),
                  1 if series_month.month == 12 else series_month.month + 1, 1)
        th_exp = last_thursday_monthly(nm, holidays)
        tu_exp = last_tuesday_monthly(nm, holidays)
        use_tuesday = (tu_exp >= CHANGEOVER_DATE)
        expiry = tu_exp if use_tuesday else th_exp
    return expiry

def format_ddMONyyyy(d: date) -> str:
    return d.strftime("%d-%b-%Y")  # match NSE OC "expiryDate" like '25-Sep-2025'

# ======= Option chain fetch =======
@dataclass
class ChainPick:
    spot: float
    atm_strike: int
    ce_ltp: Optional[float]
    pe_ltp: Optional[float]
    used_expiry_text: str  # e.g., '25-Sep-2025'

def is_index_symbol(sym: str) -> bool:
    return sym.upper() in INDEX_SYMBOLS

def pick_monthly_expiry_from_chain(chain_json: Dict, desired_ddMONyyyy: str) -> Optional[str]:
    expiries = chain_json.get("records", {}).get("expiryDates", []) or []
    for e in expiries:
        if e.strip().lower() == desired_ddMONyyyy.strip().lower():
            return e
    try:
        want = parse_date_flex(desired_ddMONyyyy)
        cands = [parse_date_flex(e) for e in expiries]
        same_month = [e for e in cands if e and want and e.year == want.year and e.month == want.month]
        if same_month:
            return format_ddMONyyyy(max(same_month))
    except Exception:
        pass
    return expiries[0] if expiries else None

def pick_monthly_expiry_from_chain_dates(expiry_dates: list, desired_ddMONyyyy: str) -> Optional[str]:
    """Pick the best matching expiry date from a list of available dates."""
    if not expiry_dates:
        return None
    
    # Try exact match first
    for e in expiry_dates:
        if e.strip().lower() == desired_ddMONyyyy.strip().lower():
            return e
    
    # Try finding same month
    try:
        want = parse_date_flex(desired_ddMONyyyy)
        cands = [parse_date_flex(e) for e in expiry_dates]
        same_month = [e for e in cands if e and want and e.year == want.year and e.month == want.month]
        if same_month:
            return format_ddMONyyyy(max(same_month))
    except Exception:
        pass
    
    # Return the closest available
    return expiry_dates[0] if expiry_dates else None

def extract_atm_ce_pe(chain_json: Dict, expiry_txt: str, symbol: str = "UNKNOWN") -> ChainPick:
    """Extract ATM CE and PE prices from option chain JSON (handles v3 API format)."""
    
    # Handle v3 API response which has 'records' with 'data' containing options
    recs = chain_json.get("records", {})
    
    # Get underlying spot price
    spot = float(recs.get("underlyingValue") or 0.0)
    if spot == 0:
        # Try to extract from data if available
        data_list = recs.get("data", [])
        for item in data_list:
            if "underlyingValue" in item:
                spot = float(item.get("underlyingValue", 0))
                if spot > 0:
                    break
    
    strikes = recs.get("strikePrices", []) or []

    if not strikes and not recs.get("data"):
        return ChainPick(spot=spot, atm_strike=0, ce_ltp=None, pe_ltp=None, used_expiry_text=expiry_txt)

    strike_prices = {}
    data_list = recs.get("data", [])
    
    for row in data_list:
        # Check if this row matches the expiry we want
        expiry_match = row.get("expiryDate") == expiry_txt or row.get("expiryDates") == expiry_txt
        if not expiry_match:
            continue
        
        strike = int(float(row.get("strikePrice", -1)))
        if strike not in strike_prices:
            strike_prices[strike] = {"ce": None, "pe": None}
        
        # Extract CE and PE prices
        if "CE" in row:
            ce_data = row["CE"]
            if isinstance(ce_data, dict) and ce_data.get("lastPrice"):
                try:
                    strike_prices[strike]["ce"] = float(ce_data["lastPrice"])
                except (ValueError, TypeError):
                    pass
        
        if "PE" in row:
            pe_data = row["PE"]
            if isinstance(pe_data, dict) and pe_data.get("lastPrice"):
                try:
                    strike_prices[strike]["pe"] = float(pe_data["lastPrice"])
                except (ValueError, TypeError):
                    pass

    valid_strikes = []
    for strike, prices in strike_prices.items():
        if prices["ce"] is not None and prices["pe"] is not None and prices["ce"] > 0 and prices["pe"] > 0:
            valid_strikes.append(strike)

    if not valid_strikes:
        # If no valid strikes, try to get closest to spot
        if strikes:
            atm_strike = min(strikes, key=lambda x: abs(x - spot))
        elif strike_prices:
            atm_strike = min(strike_prices.keys(), key=lambda x: abs(x - spot))
        else:
            atm_strike = 0
        return ChainPick(spot=spot, atm_strike=int(atm_strike), ce_ltp=None, pe_ltp=None, used_expiry_text=expiry_txt)

    # Find the best ATM strike
    spot_range = spot * 0.2
    nearby_strikes = [s for s in valid_strikes if abs(s - spot) <= spot_range]
    if not nearby_strikes:
        nearby_strikes = valid_strikes

    best_strike = None
    best_score = float('inf')
    for strike in nearby_strikes:
        ce_price = strike_prices[strike]["ce"]
        pe_price = strike_prices[strike]["pe"]
        # Prefer strikes where CE and PE are close in price
        if ce_price <= pe_price:
            score = abs(pe_price - ce_price) + abs(strike - spot) * 0.01
        else:
            score = abs(pe_price - ce_price) + 100 + abs(strike - spot) * 0.01
        if score < best_score:
            best_score = score
            best_strike = strike

    if best_strike is None:
        best_strike = min(nearby_strikes, key=lambda x: abs(x - spot))

    ce_ltp = strike_prices[best_strike]["ce"]
    pe_ltp = strike_prices[best_strike]["pe"]

    print(f"[STRIKE SELECTION] {symbol} | Spot: {spot} | Selected Strike: {best_strike} | CE: {ce_ltp} | PE: {pe_ltp} | Straddle: {ce_ltp + pe_ltp}")
    return ChainPick(spot=spot, atm_strike=int(best_strike), ce_ltp=ce_ltp, pe_ltp=pe_ltp, used_expiry_text=expiry_txt)

# ======= Ban list + MWPL (guidance only) =======
def fetch_official_ban_symbols() -> Set[str]:
    try:
        s = make_session()
        r = s.get(NSE_BAN_CSV_URL, timeout=10)
        r.raise_for_status()
        data = r.content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(data))
        out: Set[str] = set()
        for row in reader:
            if len(row) >= 2 and row[0].strip().isdigit():
                out.add(row[1].strip().upper())
        return out
    except Exception:
        return set()

def fetch_public_mwpl_snapshot() -> Dict[str, float]:
    UA = {"User-Agent": "Mozilla/5.0"}
    out: Dict[str, float] = {}

    def parse_pairs(html: str):
        for m in re.finditer(r">([A-Z0-9]{2,})</td>\s*<td[^>]*>\s*([7-9]\d(?:\.\d+)?)\s*%</td>", html, flags=re.I):
            sym = m.group(1).upper().strip()
            pct = float(m.group(2))
            out[sym] = max(out.get(sym, 0.0), pct)

    try:
        html = requests.get("https://www.5paisa.com/nse-ban-list", headers=UA, timeout=10).text
        parse_pairs(html)
    except Exception:
        pass
    try:
        html = requests.get("https://www.niftytrader.in/ban-list", headers=UA, timeout=10).text
        parse_pairs(html)
    except Exception:
        pass

    return out

# ======= Margin helpers (SPAN + Exposure) =======
def _default_margin_pct(sym: str, is_index: bool) -> Tuple[float, float]:
    """
    Return (span_pct, exposure_pct) as decimals.
    Tuned closer to NSE SPAN defaults.
    """
    if is_index:
        return 0.10, 0.03   # ~13% total for indices
    else:
        return 0.139, 0.07   # ~22% total for stock options/futures

def resolve_margin_pcts(row: pd.Series, sym: str, is_index: bool) -> Tuple[float, float]:
    span_pct = None
    exp_pct = None
    if "SPAN %" in row and pd.notnull(row["SPAN %"]):
        span_pct = float(row["SPAN %"]) / 100.0
    if "Exposure %" in row and pd.notnull(row["Exposure %"]):
        exp_pct = float(row["Exposure %"]) / 100.0
    d_span, d_exp = _default_margin_pct(sym, is_index)
    return (span_pct if span_pct is not None else d_span,
            exp_pct if exp_pct is not None else d_exp)

# SOMETIMES THE MARGIN ARE INCREASED ON MANY BASES, SO PLEASE TAKE THE CALCULATION FOR THE MARGIN IF I INTITED THE POSITION
def compute_initial_margin(contract_value: Optional[float],
                           span_pct: float,
                           exposure_pct: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if not contract_value:
        return None, None, None
    span_r = round(contract_value * span_pct, 2)
    exp_r  = round(contract_value * exposure_pct, 2)
    total  = round(span_r + exp_r, 2)
    return span_r, exp_r, total

# ======= Excel helpers =======
def estimate_margin_lakhs(_lot_size: Optional[int]) -> Optional[float]:
    # kept for backward compatibility; not used for utilisation now
    if not _lot_size:
        return None
    return round(1.4 * (_lot_size / max(_lot_size, 1)), 2)

def update_excel():
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    # normalize symbol
    df["symbol_clean"] = (
        df["STOCK NAME"].astype(str)
        .str.replace(EXCHANGE_PREFIX, "", regex=False)
        .str.strip()
        .str.upper()
    )

    banned = fetch_official_ban_symbols()
    mwpl_map = fetch_public_mwpl_snapshot()

    now = datetime.now()
    desired_expiry = monthly_expiry_for_series(now.date())
    desired_expiry_txt = format_ddMONyyyy(desired_expiry)

    updates = []
    for _, row in df.iterrows():
        sym = row.get("symbol_clean")
        if not sym or sym.lower() == "nan":
            updates.append(None)
            continue

        try:
            # First get available expiry dates for this symbol
            dates_response = fetch_oc_dates(sym, is_index_symbol(sym))
            if not dates_response or not dates_response.get("expiryDates"):
                raise RuntimeError(f"No expiry dates found for {sym}")
            
            # Extract available expiry dates
            available_expiries = dates_response.get("expiryDates", [])
            
            # Pick the best matching expiry date
            picked_expiry_txt = pick_monthly_expiry_from_chain_dates(available_expiries, desired_expiry_txt)
            if not picked_expiry_txt:
                raise RuntimeError(f"No suitable expiry found for {sym}")
            
            # Now fetch the actual option chain data for this expiry
            oc = fetch_oc_json(sym, is_index_symbol(sym), picked_expiry_txt)
            if not oc or not oc.get("records"):
                raise RuntimeError(f"Failed to fetch option chain for {sym} with expiry {picked_expiry_txt}")

            pick = extract_atm_ce_pe(oc, picked_expiry_txt, sym)

            lot_size = int(row.get("Lot Size")) if pd.notnull(row.get("Lot Size")) else None
            init_strike = row.get("Price When Straddle Initiated")
            init_total  = row.get("Straddle Total When Initiated")

            ce = pick.ce_ltp if pick.ce_ltp is not None else None
            pe = pick.pe_ltp if pick.pe_ltp is not None else None
            straddle = (ce + pe) if (ce is not None and pe is not None) else None

            # ---- NEW: Margin based on option short notional: strike * lot_size ----
            contract_value = (pick.atm_strike * lot_size) if (pick.atm_strike and lot_size) else None
            span_pct, exp_pct = resolve_margin_pcts(row, sym, is_index_symbol(sym))
            span_r, exp_r, init_margin_r = compute_initial_margin(contract_value, span_pct, exp_pct)

            # deltas vs initiated
            diff_strike_pct = round(((pick.atm_strike - init_strike) / init_strike) * 100, 2) if pd.notnull(init_strike) else None
            diff_price_pct  = round(((straddle - init_total) / init_total) * 100, 2) if (pd.notnull(init_total) and straddle) else None

            # premiums
            current_straddle_premium = round(straddle * lot_size, 2) if (straddle and lot_size) else None
            executed_straddle_premium = round(init_total * lot_size, 2) if (pd.notnull(init_total) and lot_size) else None
            net_diff_premium = round(current_straddle_premium - executed_straddle_premium, 2) if (current_straddle_premium and executed_straddle_premium) else None

            # utilisation now uses *computed* initial margin
            margin_utilised_pct = round((current_straddle_premium / init_margin_r) * 100, 2) if (current_straddle_premium and init_margin_r) else None

            # range %
            range_pct = round(abs((pick.atm_strike - pick.spot) / pick.spot) * 100, 2) if pick.spot else None

            # Console print
            timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
            print(f"Time Stamp: {timestamp}")
            print(f"Name: {sym}")
            print(f"Live Price: {pick.spot}")
            print(f"Current Straddle Strike: {pick.atm_strike}")
            print(f"Current Straddle Price: {round(straddle, 2) if straddle else 'N/A'}")
            print(f"Lot Size: {lot_size}")
            print(f"SPAN % / Exposure %: {round(span_pct*100,2)}% / {round(exp_pct*100,2)}%")
            print(f"SPAN (₹): {span_r if span_r else 'N/A'} | Exposure (₹): {exp_r if exp_r else 'N/A'} | Initial Margin (₹): {init_margin_r if init_margin_r else 'N/A'}")
            print(f"Current Straddle Premium: {current_straddle_premium if current_straddle_premium else 'N/A'}")
            print(f"Executed Straddle Strike: {init_strike if pd.notnull(init_strike) else 'N/A'}")
            print(f"Executed Straddle Price: {init_total if pd.notnull(init_total) else 'N/A'}")
            print(f"Executed Straddle Premium: {executed_straddle_premium if executed_straddle_premium else 'N/A'}")
            print(f"Net Diff in Premium: {net_diff_premium if net_diff_premium else 'N/A'}")
            print(f"% of Margin Utilised: {margin_utilised_pct}%" if margin_utilised_pct else "% of Margin Utilised: N/A")
            print(f"Range %: {range_pct}%" if range_pct else "Range %: N/A")
            print("-" * 70)

            updates.append({
                "Live Price": round(pick.spot, 2) if pick.spot else None,
                "CURRENT STRADDLE STRIKE": pick.atm_strike,
                "CURRENT STRADDLE PRICE": round(straddle, 2) if straddle else None,
                "Net Premium Credited": round(straddle * lot_size, 2) if (straddle and lot_size) else None,

                # NEW margin outputs
                "SPAN (₹)": span_r,
                "Exposure (₹)": exp_r,
                "Initial Margin (₹)": init_margin_r,
                "SPAN %": round(span_pct * 100, 2) if span_pct is not None else None,
                "Exposure %": round(exp_pct * 100, 2) if exp_pct is not None else None,

                # legacy field for quick glance
                "Margin Needed (In Lakhs)": round((init_margin_r / 100000), 2) if init_margin_r else None,

                "DIFF BETWEEN INITIATED AND CURRENT STRIKE(%)": diff_strike_pct,
                "DIFF BETWEEN INITIATED AND CURRENT PRICE(%)": diff_price_pct,
                "Last Updated": now.strftime("%Y-%m-%d %H:%M:%S"),
                "Expiry Used": picked_expiry_txt,
                "Ban Status": "BANNED" if sym in banned else "OK",
                "MWPL % (public)": mwpl_map.get(sym)
            })
        except Exception as e:
            updates.append({
                "Live Price": None,
                "CURRENT STRADDLE STRIKE": None,
                "CURRENT STRADDLE PRICE": None,
                "Net Premium Credited": None,
                "SPAN (₹)": None,
                "Exposure (₹)": None,
                "Initial Margin (₹)": None,
                "SPAN %": None,
                "Exposure %": None,
                "Margin Needed (In Lakhs)": None,
                "DIFF BETWEEN INITIATED AND CURRENT STRIKE(%)": None,
                "DIFF BETWEEN INITIATED AND CURRENT PRICE(%)": None,
                "Last Updated": f"ERROR: {str(e)[:50]}",
                "Expiry Used": desired_expiry_txt,
                "Ban Status": "ERROR",
                "MWPL % (public)": None
            })
            print(f"[UPDATE ERROR] {sym}: {str(e)[:80]}")
        
        # Small delay between symbols to avoid overwhelming NSE
        time.sleep(0.3)

    upd = pd.DataFrame(updates)
    df.loc[:, upd.columns] = upd.values
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="w") as w:
        df.drop(columns=["symbol_clean"], errors="ignore").to_excel(w, SHEET_NAME, index=False)

    # Color rows after write
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    headers = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    col_ban  = headers.get("Ban Status")
    col_mwpl = headers.get("MWPL % (public)")

    for r in range(2, ws.max_row + 1):
        fill = None
        ban = ws.cell(row=r, column=col_ban).value if col_ban else None
        mw  = ws.cell(row=r, column=col_mwpl).value if col_mwpl else None
        if ban == "BANNED":
            fill = FILL_RED
        else:
            try:
                pct = float(mw) if mw not in (None, "", "None") else None
            except Exception:
                pct = None
            if pct is not None:
                if pct >= 90.0:
                    fill = FILL_ORANGE
                elif pct >= 80.0:
                    fill = FILL_YELLOW
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    # Legend sheet
    legend = wb["Legend"] if "Legend" in wb.sheetnames else wb.create_sheet("Legend")
    legend.delete_rows(1, legend.max_row)
    legend.append(["Color", "Meaning"])
    legend.append(["Red", "Official NSE Ban List (Security in ban period)"])
    legend.append(["Orange", "High MWPL ≥ 90% (public trackers) – likely to be banned"])
    legend.append(["Yellow", "Elevated MWPL 80–90% (public trackers) – watchlist"])
    legend.append(["Note", "MWPL% uses public pages for guidance; rely on NSE official ban CSV for compliance."])
    legend.append(["Initial Margin", "Computed as SPAN (₹) + Exposure (₹). % can be overridden per row via 'SPAN %' and 'Exposure %'."])
    wb.save(EXCEL_PATH)

# ======= entrypoint =======
if __name__ == "__main__":
    # DEBUG: Test with a single symbol first
    print("=" * 70)
    print("TESTING NSE API WITH SINGLE SYMBOL")
    print("=" * 70)
    try:
        # First get available dates for RELIANCE
        test_dates = fetch_oc_dates("RELIANCE", is_index=False)
        print(f"[TEST DATES] Got response: {bool(test_dates)}")
        if test_dates and test_dates.get("expiryDates"):
            print(f"[TEST DATES] Available expiries: {test_dates['expiryDates'][:3]}...")
            # Now fetch data for the first available expiry
            first_expiry = test_dates["expiryDates"][0]
            test_oc = fetch_oc_json("RELIANCE", is_index=False, expiry_date=first_expiry)
            print(f"[TEST RESULT] Got data response: {bool(test_oc)}")
            print(f"[TEST RESULT] Response keys: {list(test_oc.keys())}")
            if test_oc.get('records'):
                print(f"[TEST RESULT] Records keys: {list(test_oc['records'].keys())}")
        else:
            print("[TEST ERROR] Failed to get expiry dates")
    except Exception as e:
        print(f"[TEST ERROR] {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 70)
    print("RUNNING FULL UPDATE")
    print("=" * 70)
    update_excel()
    print("Updated:", EXCEL_PATH)
